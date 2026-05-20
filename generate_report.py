"""
训练结果汇总：从 history.json 生成 Excel 报表。
用法：
  conda activate diet-training
  python generate_report.py runs/classify/food_cls_lora

会在 runs/classify/food_cls_lora/ 下生成 training_report.xlsx
"""

import json, sys, os
from pathlib import Path
try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl -q")
    import openpyxl


def generate_report(run_dir: str) -> None:
    run_path = Path(run_dir)
    history_path = run_path / "history.json"
    if not history_path.exists():
        print(f"[ERROR] 未找到 {history_path}")
        return

    with open(history_path) as f:
        history = json.load(f)

    if not history:
        print("[ERROR] history.json 为空")
        return

    wb = openpyxl.Workbook()

    # ── Sheet 1: 汇总 ──
    ws_summary = wb.active
    ws_summary.title = "汇总"

    best_epoch = max(history, key=lambda x: x["val_acc"])
    last = history[-1]

    summary_data = [
        ("总训练轮数", len(history)),
        ("最佳验证准确率 (val_acc)", f"{best_epoch['val_acc']:.4f}"),
        ("最佳轮次", best_epoch["epoch"]),
        ("最终训练损失", f"{last['train_loss']:.4f}"),
        ("最终验证损失", f"{last['val_loss']:.4f}"),
        ("最终训练准确率", f"{last['train_acc']:.4f}"),
        ("最终验证准确率", f"{last['val_acc']:.4f}"),
    ]

    for i, (key, val) in enumerate(summary_data, 1):
        ws_summary.cell(row=i, column=1, value=key)
        ws_summary.cell(row=i, column=2, value=val)

    # ── Sheet 2: 每轮详情 ──
    ws_detail = wb.create_sheet("每轮详情")
    headers = ["epoch", "train_loss", "train_acc", "val_loss", "val_acc", "lr"]
    for col, header in enumerate(headers, 1):
        ws_detail.cell(row=1, column=col, value=header)

    for row_idx, record in enumerate(history, 2):
        for col_idx, key in enumerate(headers, 1):
            ws_detail.cell(row=row_idx, column=col_idx, value=record.get(key, ""))

    # 列宽自动调整
    for ws in [ws_summary, ws_detail]:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output_path = run_path / "training_report.xlsx"
    wb.save(str(output_path))
    print(f"[OK] 报表已生成: {output_path}")


if __name__ == "__main__":
    run_dir = sys.argv[1] if len(sys.argv) > 1 else "runs/classify/food_cls_lora"
    generate_report(run_dir)
